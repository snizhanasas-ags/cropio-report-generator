import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from copy import copy
from io import BytesIO


st.set_page_config(
    page_title="Cropio Report Generator",
    page_icon="🚜",
    layout="wide"
)


st.title("🚜 Автоматичне формування звіту Cropio")

st.write(
    """
    Завантажте:
    1. Звіт роботи техніки Cropio
    2. Звіт видачі палива Cropio
    3. Ваш Excel-шаблон готового звіту
    """
)


work_file = st.file_uploader(
    "📄 1. Звіт роботи техніки Cropio",
    type=["xlsx"]
)


fuel_file = st.file_uploader(
    "⛽ 2. Звіт видачі палива Cropio",
    type=["xlsx"]
)


template_file = st.file_uploader(
    "📑 3. Готовий шаблон звіту",
    type=["xlsx"]
)
def find_column(df, variants):
    """
    Пошук колонки за ключовими словами
    """
    for col in df.columns:
        name = str(col).lower().strip()

        for variant in variants:
            if variant.lower() in name:
                return col

    return None


def read_cropio_work(file):
    """
    Читання звіту роботи техніки
    """

    excel = pd.ExcelFile(file)

    # Беремо перший лист, якщо структура стандартна
    sheet = excel.sheet_names[0]

    df = pd.read_excel(
        file,
        sheet_name=sheet
    )


    machine = find_column(
        df,
        ["машина", "техніка", "vehicle", "machine"]
    )

    if not machine:
        raise Exception(
            "Не знайдено колонку техніки у звіті роботи"
        )


    return df, machine



def read_cropio_fuel(file):
    """
    Читання звіту палива
    """

    excel = pd.ExcelFile(file)

    sheet = excel.sheet_names[0]


    # пробуємо різні рядки заголовків
    for header in range(0, 15):

        df = pd.read_excel(
            file,
            sheet_name=sheet,
            header=header
        )

        cols = " ".join(
            [str(x).lower() for x in df.columns]
        )


        if (
            "отрим" in cols
            or "машин" in cols
            or "технік" in cols
        ):
            break


    machine = find_column(
        df,
        [
            "отримувач",
            "машина",
            "техніка",
            "machine"
        ]
    )


    amount = find_column(
        df,
        [
            "кількість",
            "літр",
            "паливо",
            "amount"
        ]
    )


    source = find_column(
        df,
        [
            "джерело",
            "азс",
            "місце",
            "заправ"
        ]
    )


    if not machine:
        raise Exception(
            "Не знайдено техніку у звіті палива"
        )


    return df, machine, amount, source
    def prepare_fuel_data(df, machine_col, amount_col, source_col):
    """
    Підготовка палива:
    - сума заправок;
    - розподіл по джерелах
    """

    if amount_col is None:
        raise Exception(
            "Не знайдено колонку кількості палива"
        )


    df = df.copy()


    # перетворення літрів у число
    df[amount_col] = (
        df[amount_col]
        .astype(str)
        .str.replace(",", ".")
    )

    df[amount_col] = pd.to_numeric(
        df[amount_col],
        errors="coerce"
    )


    df = df.dropna(
        subset=[machine_col, amount_col]
    )


    # якщо немає джерела
    if source_col is None:
        df["Джерело"] = "Паливо"
        source_col = "Джерело"



    fuel = (
        df.groupby(
            [
                machine_col,
                source_col
            ]
        )[amount_col]
        .sum()
        .reset_index()
    )


    return fuel



def get_machine_fuel(
        fuel_df,
        machine,
        machine_col,
        source_col,
        amount_col
):
    """
    Отримати паливо для конкретної техніки
    """

    result = fuel_df[
        fuel_df[machine_col].astype(str)
        ==
        str(machine)
    ]


    fuel_dict = {}


    for _, row in result.iterrows():

        source = str(
            row[source_col]
        )

        amount = row[amount_col]


        fuel_dict[source] = amount


    return fuel_dict



def merge_work_and_fuel(
        work_df,
        work_machine,
        fuel_df,
        fuel_machine,
        fuel_source,
        fuel_amount
):
    """
    Додаємо паливо тільки у перший рядок машини
    """

    result = work_df.copy()


    result["__паливо__"] = None


    used_machines = set()


    for index, row in result.iterrows():

        machine = row[work_machine]


        if machine in used_machines:
            continue


        fuel = get_machine_fuel(
            fuel_df,
            machine,
            fuel_machine,
            fuel_source,
            fuel_amount
        )


        if fuel:

            result.at[
                index,
                "__паливо__"
            ] = fuel


        used_machines.add(machine)



    # техніка тільки з палива
    work_machines = set(
        result[work_machine]
        .astype(str)
    )


    fuel_machines = set(
        fuel_df[fuel_machine]
        .astype(str)
    )


    only_fuel = fuel_machines - work_machines


    for machine in only_fuel:

        new_row = {
            col: ""
            for col in result.columns
        }


        new_row[work_machine] = machine


        new_row["__паливо__"] = get_machine_fuel(
            fuel_df,
            machine,
            fuel_machine,
            fuel_source,
            fuel_amount
        )


        result.loc[
            len(result)
        ] = new_row



    return result
def copy_row_style(ws, source_row, target_row):
    """
    Копіювання оформлення рядка шаблону
    """

    for col in range(1, ws.max_column + 1):

        source = ws.cell(
            source_row,
            col
        )

        target = ws.cell(
            target_row,
            col
        )

        if source.has_style:
            target._style = copy(
                source._style
            )

        if source.number_format:
            target.number_format = (
                source.number_format
            )

        if source.alignment:
            target.alignment = copy(
                source.alignment
            )

        if source.border:
            target.border = copy(
                source.border
            )

        if source.fill:
            target.fill = copy(
                source.fill
            )

        if source.font:
            target.font = copy(
                source.font
            )


def write_to_template(
        template_file,
        data,
        machine_col
):

    wb = load_workbook(
        template_file
    )

    ws = wb.active


    # рядок початку таблиці
    start_row = 2


    # очищення старих даних
    for row in ws.iter_rows(
        min_row=start_row
    ):
        for cell in row:
            cell.value = None


    # беремо стиль останнього рядка шаблону
    style_row = ws.max_row


    current_row = start_row


    for _, row in data.iterrows():

        if current_row > ws.max_row:

            ws.insert_rows(
                current_row
            )


        copy_row_style(
            ws,
            style_row,
            current_row
        )


        col_index = 1


        for value in row:

            if col_index <= ws.max_column:

                cell = ws.cell(
                    current_row,
                    col_index
                )


                cell.value = value


            col_index += 1


        # паливо
        fuel = row.get(
            "__паливо__"
        )


        if isinstance(
            fuel,
            dict
        ):

            for source, amount in fuel.items():

                # шукаємо колонку АЗС
                for cell in ws[current_row]:

                    if (
                        str(cell.value)
                        ==
                        str(source)
                    ):

                        fuel_cell = ws.cell(
                            current_row,
                            cell.column + 1
                        )

                        fuel_cell.value = amount


                        # червоний колір
                        fuel_cell.font = copy(
                            fuel_cell.font
                        )

                        fuel_cell.font = Font(
                            name=fuel_cell.font.name,
                            size=fuel_cell.font.size,
                            bold=fuel_cell.font.bold,
                            color="FF0000"
                        )


        current_row += 1


    output = BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    return output
    if st.button("🟢 Сформувати звіт"):

    try:

        work, work_machine = read_cropio_work(
            work_file
        )


        fuel, fuel_machine, fuel_amount, fuel_source = read_cropio_fuel(
            fuel_file
        )


        fuel_ready = prepare_fuel_data(
            fuel,
            fuel_machine,
            fuel_amount,
            fuel_source
        )


        result = merge_work_and_fuel(
            work,
            work_machine,
            fuel_ready,
            fuel_machine,
            fuel_source,
            fuel_amount
        )


        file = write_to_template(
            template_file,
            result,
            work_machine
        )


        st.success(
            "✅ Звіт успішно сформовано"
        )


        st.download_button(
            "⬇️ Завантажити готовий Excel",
            file,
            "Cropio_final_report.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    except Exception as e:

        st.error(
            f"Помилка: {e}"
        )
